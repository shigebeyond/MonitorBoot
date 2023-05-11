import math
import os
import re

import pandas as pd
from openpyxl.utils import get_column_letter
from pyutilb import ts
from pyutilb.file import *
from pyutilb.log import log
from pyutilb.strs import substr_before
from pyutilb.tail import Tail
from pyutilb.util import set_vars, val2df
from ExcelBoot.boot import Boot as EBoot

# ExcelBoot的步骤文件
gcs_excel_boot_yaml = __file__.replace("gc_log_parser.py", "eb-gcs2xlsx.yml")
compare_gcs_excel_boot_yaml = __file__.replace("gc_log_parser.py", "eb-compare_gcs2xlsx.yml")

'''
gc日志解析，主要用于识别频繁gc 或 gc效率不高
    -Xms20M -Xmx20M -Xmn10M -XX:SurvivorRatio=8 # 堆大小
    -verbose:gc -XX:+PrintGCDetails  -Xloggc:gc.log # gc日志 
    -XX:+HeapDumpOnOutOfMemoryError # oom时dump出堆快照，如java_pid30949.hprof
    -XX:MaxTenuringThreshold=1 -XX:+PrintTenuringDistribution # 晋升老年代的age阈值 
'''
class GcLogParser(object):

    def __init__(self, log_file):
        self.log_file = os.path.abspath(log_file)
        self.start_time = os.path.getctime(self.log_file)
        self.gcs = [] # 收集日志信息

    # 获得上一条年轻代或老年代的gc信息
    def last_gc(self, is_full):
        is_full = bool(is_full)
        for gc in reversed(self.gcs):
            if gc['is_full'] == is_full:
                return gc

        return None

    # 解析gc日志
    # :param log_file 日志文件路径
    def parse(self):
        txt = read_file(self.log_file)
        for line in txt.splitlines():
            # 解析单行
            gc = self.parse_gc_line(line)

    # 快速判断是否gc line
    def is_gc_line(self, line):
        return ': [' in line \
               and '->' in line \
               and 'secs] [Times: user=' in line

    def parse_gc_line(self, line):
        '''
        解析gc日志行, 其中 fullgc比gc多了[Perm:28671K->28635K(28672K)]
           Parallel Scavenge GC
            0.084: [GC (Allocation Failure) [PSYoungGen: 1525K->512K(1536K)] 3556K->2886K(5632K), 0.0039928 secs] [Times: user=0.01 sys=0.00, real=0.00 secs]
            0.089: [Full GC (Ergonomics) [PSYoungGen: 1536K->0K(1536K)] [ParOldGen: 3312K->4088K(4096K)] 4848K->4088K(5632K), [Metaspace: 3313K->3313K(1056768K)], 0.0416957 secs] [Times: user=0.13 sys=0.00, real=0.04 secs]
          ParNew/CParNew/MS GC，与 Parallel Scavenge GC相比，其jvm time(如0.064: )重复了2次，需要特殊处理
            0.064: [GC (Allocation Failure) 0.064: [ParNew: 509K->64K(576K), 0.0032549 secs] 509K->282K(1984K), 0.0033544 secs] [Times: user=0.01 sys=0.00, real=0.00 secs]
            104429.457: [Full GC (System) 104429.457: [CMS: 219741K->215266K(1835008K), 0.5469450 secs] 244623K->215266K(2070976K), [CMS Perm : 128846K->128831K(262144K)], 0.5470720 secs] [Times: user=0.54 sys=0.00, real=0.55 secs]
        :param line:
        :return:
        '''
        if not self.is_gc_line(line):
            return None
        try:
            gens = []
            # print("解析gc行: " + line)
            line = line.replace('--', '') # 特殊例子: 0.095: [GC (Allocation Failure) --[PSYoungGen: 1520K->1520K(1536K)] 4784K->5608K(5632K), 0.0099458 secs] [Times: user=0.03 sys=0.00, real=0.01 secs]
            # 1 处理几个年代的空间+时间
            items = re.findall('\[[^\[^\]]+\]', line)
            for item in items:
                # 1.1 解析单个年代
                if item.startswith('[Times:'): # 忽略 [Times: user=0.13 sys=0.00, real=0.04 secs]
                    # print("忽略非年代: " + item)
                    pass
                else: # 解析单个年代
                    data = self.parse_gen(item)
                    if data is None:
                        raise Exception("解析年代失败: " + item)
                    # print("解析年代: " + item + ", 结果为: " + str(data))
                    gens.append(data) # 记录解析结果
                # 1.2 去掉解析过的年代部分字符串
                line = line.replace(item, '')
                # print("剩余gc行: " + line)

            # 2 处理总的空间+时间
            # 如 0.089: [Full GC (Ergonomics)   4848K->4088K(5632K), , 0.0416957 secs]
            # 如 0.084: [GC (Allocation Failure)  3556K->2886K(5632K), 0.0039928 secs]
            # 如 0.064: [GC (Allocation Failure) 0.064:  509K->282K(1984K), 0.0033544 secs] -- ParNew/CMS GC与 Parallel Scavenge GC相比，其jvm time(如0.064: )重复了2次，需要特殊处理；注: 也可能不重复，但两个时间相差0.001秒，反正就是同一个gc段内
            # 优先处理 jvm time： 干掉
            jvm_time = substr_before(line, ': [') # gc发生时vm运行了多少秒
            line = line.replace(jvm_time + ': ', '')  # 干掉 jvm time，有可能2次
            if ':' in jvm_time: # 前面有可能有时间: `time: jvm_time: [`, 如 2019-03-28T18:09:15.774+0800: 389.142: [
                time, jvm_time = jvm_time.rsplit(':', 1)
                # todo: 也解析time(系统时间)
            # 规整line为年代格式
            line = line.replace(', ,', ',').replace(') ', '):')
            # 解析总的年代
            gc = self.parse_gen(line)
            gc['jvm_time'] = float(jvm_time)  # gc发生时vm运行了多少秒
            is_full = 'Full GC' in line
            # 计算两次gc之间的时间间隔
            lastgc = self.last_gc(is_full)
            if lastgc is None:
                # gc['interval'] = gc['jvm_time'] # 你不知道他是从啥时开始监控日志的，也不知道监控之前有没有gc过
                gc['interval'] = 0
            else:
                gc['interval'] = gc['jvm_time'] - lastgc['jvm_time']
            gc['is_full'] = is_full
            # print("解析总年代: " + line + ", 结果为: " + str(gc))

            # 3 展平多个年代
            # gc['gens'] = gens
            self.flatten_gens(gc, gens)

            # print(gc)
            self.gcs.append(gc)
            return gc
        except Exception as ex:
            log.error("GcLogParser.parse_gc_line()异常: " + str(ex), exc_info=ex)
            return None

    # 解析一代: 如 [PSYoungGen: 1536K->0K(1536K)]
    def parse_gen(self, gen):
        mat = re.search('\[([^:]+): +(\d+)K->(\d+)K\((\d+)K\)(, ([\d\.]+) secs)?', gen)
        if mat == None:
            return None
        data = {
            'name': mat.group(1),
            'before': float(mat.group(2)),
            'after': float(mat.group(3)),
            'total': float(mat.group(4)),
        }
        # 总的才有secs部分
        secs = mat.group(6)
        if secs is not None:
            data['costtime'] = float(secs)
        return data

    def flatten_gens(self, gc, gens):
        '''
        展平一次gc中的多个年代的变更：二维变为一维
           二维: 年代的数组  [{'name': 'PSYoungGen', 'before': 1536.0, 'after': 0.0, 'total': 1536.0}, {'name': 'ParOldGen', 'before': 3312.0, 'after': 4088.0, 'total': 4096.0}, {'name': 'Metaspace', 'before': 3313.0, 'after': 3313.0, 'total': 1056768.0}]
           一维: 年代名作为key的前缀 {'PSYoungGen.before': 1536.0, 'PSYoungGen.after': 0.0, 'PSYoungGen.total': 1536.0, 'ParOldGen.before': 3312.0, 'ParOldGen.after': 4088.0, 'ParOldGen.total': 4096.0, 'Metaspace.before': 3313.0, 'Metaspace.after': 3313.0, 'Metaspace.total': 1056768.0}
        :param gc:
        :param gens:
        :return:
        '''
        for gen in gens:
            name = gen['name']
            del gen['name']
            for k, v in gen.items():
                key = name + '.' + k
                gc[key] = v

    # 获得所有 gc
    def all_gcs(self):
        return pd.DataFrame(self.gcs, columns=['name', 'before', 'after', 'total', 'costtime', 'jvm_time', 'interval'])

    def filter_gcs(self, is_full):
        # return [gc for gc in self.gcs if gc['is_full'] == is_full]
        ret = []
        for gc in self.gcs:
            if gc['is_full'] == is_full:
                gc2 = gc.copy()
                del gc2['is_full']
                ret.append(gc2)
        return ret

    # 获得full gc
    def full_gcs(self):
        return self.filter_gcs(True)

    # 获得minor gc
    def minor_gcs(self):
        return self.filter_gcs(False)

    def gcs2xlsx(self, filename_pref, bins=None, interval=None):
        '''
        导出gc信息
        :param filename_pref:
        :param bins: 分区数，bins与interval参数是二选一
        :param interval: 分区的时间间隔，单位秒，bins与interval参数是二选一
        :return:
        '''
        if len(self.gcs) == 0:
            log.warning("无gc记录可导出")
            return

        # excel文件名
        filename_pref = filename_pref or 'JvmGC'
        now = ts.now2str("%Y%m%d%H%M%S")
        file = f'{filename_pref}-{now}.xlsx'
        # 设置变量
        all_gcs = self.all_gcs()
        minor_gcs = self.minor_gcs()
        full_gcs = self.full_gcs()

        vars = {
            'file': file,
            # gc记录
            'all_gcs': all_gcs,
            'minor_gcs': minor_gcs,
            'full_gcs': full_gcs,
            # 将gc记录按时间划分区间，并统计各区间个数+耗时
            'all_gc_bins': self.gcs2bins(all_gcs, bins=bins, interval=interval),
            'full_gc_bins': self.gcs2bins(full_gcs, bins=bins, interval=interval),
            'minor_gc_bins': self.gcs2bins(minor_gcs, bins=bins, interval=interval),
        }
        set_vars(vars)

        # 导出excel
        boot = EBoot()
        boot.run_1file(gcs_excel_boot_yaml)
        return file

    @classmethod
    def gcs2bins(self, gcs, bins=None, interval=None):
        '''
        将gc记录按时间划分区间，并统计各区间个数+耗时
        :param gcs: gc的df
        :param bins: 分区个数
        :param interval: 时间间隔，如果要对比2次gc的频率或耗时，特别是优化前后，需在同一个时间粒度(interval)上对比
        :return:
        '''
        df = val2df(gcs)
        if bins is None:
            if interval is None:
                # raise Exception("未指定参数：bins 或 interval")
                bins = min(8, len(gcs))
            else:
                bins = math.ceil(df['jvm_time'].max() / interval)  # 区间的right不是interval的倍数(如interval=2，区间right应该是2/4/6，但生成的却是1.8,3.9,5.9之类的)，不易读，必须手动拼接区间
                bins = [interval * i for i in range(0, bins + 1)]  # 手动拼接区间
        bins = pd.cut(df['jvm_time'], bins=bins, include_lowest=True, right=True, duplicates='drop')
        # Series: index是区间Interval，如[0.054, 4.428), [4.428, 8.802), [8.802, 13.189)
        counts = pd.value_counts(bins)  # 统计各区间的个数

        # 统计各区间的耗时
        costtimes = []
        for bin in counts.index: # counts.index不是有序的
            costime = 0
            for i, r in df.iterrows():
                if r['jvm_time'] in bin:
                    costime += r['costtime']
                    break
            costtimes.append(costime)

        # 返回
        df2 = pd.DataFrame([], columns=['time', 'bin', 'count', 'costtime'])
        df2['bin'] = counts.index  # 区间
        df2['bin'] = df2['bin'].astype('str')
        df2['time'] = [bin.right for bin in counts.index]  # 区间
        df2['count'] = counts.to_list()  # 各区间的个数
        df2['costtime'] = costtimes  # 各区间的耗时
        return df2.sort_values(by='time')  # 按时间排序

    @classmethod
    def compare_gclogs2xlsx(cls, logs, interval, filename_pref = None):
        '''
        对比多个gc log，并将对比结果存到excel中
        :param logs gc log
        :param interval: 分区的时间间隔，单位秒，必填，2个gc log的对比必须基于同一个时间维度与粒度
        :param filename_pref:
        :return:
        '''
        # 1 修正参数
        if logs is None or len(logs) == 0:
            log.warning("无gc log可对比")
            return
        if not isinstance(logs, (list, dict)):
            raise Exception('logs参数期望是list或dict类型，但真实是' + type(logs))
        # list转dict
        if isinstance(logs, list):
            # 检查文件名是否有重复
            keys = [os.path.basename(file).replace('.log', '') for file in logs]
            if len(keys) == len(set(keys)): # 如果文件名不重复，则使用文件名作为key
                logs = dict(zip(keys, logs))
            else: # 否则，使用log1/log2...作为key
                logs2 = {}
                for i in range(0, len(logs)):
                    key = 'log' + str(i + 1)
                    logs2[key] = logs[i]
                logs = logs2

        # 2 解析gc log
        log2gcs = []
        jvm_times = []
        for key, file in logs.items():
            # 解析gc log
            parser = GcLogParser(file)
            parser.parse()
            gcs = parser.all_gcs()
            # 记录
            item = {
                'key': key,
                'gcs': gcs
            }
            log2gcs.append(item)
            jvm_times.append(gcs['jvm_time'].max())
        # 3 逐个log分区
        # 根据最大的jvm time，来计算统一的时间区间
        max_jvm_time = max(jvm_times)
        bins = math.ceil(max_jvm_time / interval)  # 区间的right不是interval的倍数(如interval=2，区间right应该是2/4/6，但生成的却是1.8,3.9,5.9之类的)，不易读，必须手动拼接区间
        bins = [interval * i for i in range(0, bins + 1)]  # 手动拼接区间
        # 将gc记录按时间划分区间，并统计各区间个数+耗时
        for item in log2gcs:
            item['gc_bins'] = cls.gcs2bins(item['gcs'], bins=bins)
        # 4 合并所有log的分区数据
        # 将所有log的bins合并为一个，每个log作为一列
        # 4.1 对比不同log的分区gc计数
        compare_count_df = pd.DataFrame([], columns=['time'])
        compare_count_df['time'] = log2gcs[0]['gc_bins']['time']  # 时间
        for item in log2gcs:
            key = item['key'] + '.'
            compare_count_df[key + 'count'] = item['gc_bins']['count']
        # 4.2 对比不同log的分区gc耗时
        compare_costtime_df = pd.DataFrame([], columns=['time'])
        compare_costtime_df['time'] = log2gcs[0]['gc_bins']['time']  # 时间
        for item in log2gcs:
            key = item['key'] + '.'
            compare_costtime_df[key + 'costtime'] = item['gc_bins']['costtime']

        # 5 导出
        # excel文件名
        filename_pref = filename_pref or 'JvmGCCompare'
        now = ts.now2str("%Y%m%d%H%M%S")
        file = f'{filename_pref}-{now}.xlsx'
        # 设置变量
        vars = {
            'file': file,
            'log2gcs': log2gcs,
            'compare_count_df': compare_count_df,
            'compare_costtime_df': compare_costtime_df,
            'max_col': get_column_letter(len(log2gcs)+1), # 最大列名，用于设置列样式
            'plot_col': get_column_letter(len(log2gcs)+3) # 插入plot绘图的列
        }
        set_vars(vars)

        # 导出excel
        boot = EBoot()
        boot.run_1file(compare_gcs_excel_boot_yaml)
        return file

if __name__ == '__main__':
    # file = '../logs/gc2.log'
    file = '/home/shi/code/testing/kt-test/gc.log'
    parser = GcLogParser(file)

    # 0 生成gc log
    '''
    for i in range(1, 30):
        line = f"{i}: [Full GC (Ergonomics) [PSYoungGen: 1024K->0K(1536K)] [ParOldGen: 4070K->392K(4096K)] 5094K->392K(5632K), [Metaspace: 3332K->3332K(1056768K)], {i} secs] [Times: user=0.00 sys=0.01, real=0.01 secs]"
        print(line)
    '''

    # 1 测试解析单行
    #  Parallel Scavenge GC日志
    # line = '0.084: [GC (Allocation Failure) [PSYoungGen: 1525K->512K(1536K)] 3556K->2886K(5632K), 0.0039928 secs] [Times: user=0.01 sys=0.00, real=0.00 secs]'
    # line = '0.089: [Full GC (Ergonomics) [PSYoungGen: 1536K->0K(1536K)] [ParOldGen: 3312K->4088K(4096K)] 4848K->4088K(5632K), [Metaspace: 3313K->3313K(1056768K)], 0.0416957 secs] [Times: user=0.13 sys=0.00, real=0.04 secs]'
    # CMS GC日志
    # line = '0.064: [GC (Allocation Failure) 0.064: [ParNew: 509K->64K(576K), 0.0032549 secs] 509K->282K(1984K), 0.0033544 secs] [Times: user=0.01 sys=0.00, real=0.00 secs]'
    # line = '104429.457: [Full GC (System) 104429.457: [CMS: 219741K->215266K(1835008K), 0.5469450 secs] 244623K->215266K(2070976K), [CMS Perm : 128846K->128831K(262144K)], 0.5470720 secs] [Times: user=0.54 sys=0.00, real=0.55 secs]'
    line = '2019-03-28T18:09:15.774+0800: 389.142: [Full GC (Ergonomics) [PSYoungGen: 17010K->0K(925184K)] [ParOldGen: 2098093K->2103707K(2776064K)] 2115103K->2103707K(3701248K), [Metaspace: 62299K->62299K(1105920K)], 5.5291426 secs] [Times: user=14.83 sys=0.09, real=5.53 secs]'
    gc = parser.parse_gc_line(line)
    print(gc)

    # 2 测试解析整个log文件
    '''
    # 解析文件
    parser.parse()
    for gc in parser.gcs:
        print(gc)
    # 导出excel
    parser.gcs2xlsx(None, interval=10)
    '''

    # 3 测试tail
    '''
    t = Tail(file, from_end=False)
    i = 0
    def handle_line(line):
        gc = parser.parse_gc_line(line)
        print(gc)
        global i
        i += 1
        if i == 13:
            parser.gcs2xlsx(None)
    t.follow(handle_line)
    '''

    # 4 对比gc log
    '''
    files = [
        '/home/shi/code/testing/kt-test/gc2.log',
        '/home/shi/code/testing/kt-test/gc5.log',
    ]
    xlsx = GcLogParser.compare_gclogs2xlsx(files, interval=30)
    print(xlsx)
    '''


